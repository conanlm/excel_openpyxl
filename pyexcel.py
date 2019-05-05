# pip install openpyxl

import time
from openpyxl import Workbook
from openpyxl import load_workbook
import logging
import time
import traceback
import sys
import os.path
import os

start = time.perf_counter()
# 实例化
# wb = Workbook()
# # 激活 worksheet
# # ws = wb.active

# wb2 = load_workbook('分平台营业数据 (34).xlsx')

# ws2=wb2.worksheets[0]
# print(ws2["C2"].value)
# wb3 = load_workbook('评论率 (33).xlsx')
# ws3=wb3.worksheets[0]
# ws3["G2"].value=ws2["C2"].value
# wb3.save('评论率 (33).xlsx')

a = []
for line in open("name.txt"):
    a.append(line.strip())


def update(a):
    try:
        wb1 = load_workbook(a[0], data_only=True)
        # wb2 = load_workbook(a[2])
        wb3 = load_workbook(a[1], data_only=True)
        wb4 = load_workbook(a[5], data_only=True)
    except FileNotFoundError:
        print("二孩子，文件名写对了吗")
        input()
    else:

        try:
            arr5 = []
            for rowNum in range(3, 33):
                arr5.append(wb4.worksheets[0].cell(rowNum,4).value)
            arr1 = []
            arr2 = []
            arr3 = []
            arr4 = []
            for rowNum1 in range(2, 32):                
                arr11=[]
                for columnNum1 in range(3, 6):
                    arr11.append(wb1.worksheets[0].cell(rowNum1,columnNum1).value)            
                arr12=[]
                for columnNum2 in range(11, 23):
                    arr12.append(wb1.worksheets[0].cell(rowNum1,columnNum2).value)
                arr13=[]
                for columnNum3 in range(28, 38):
                    arr13.append(wb1.worksheets[0].cell(rowNum1,columnNum3).value)
                arr41=[]
                for columnNum4 in range(3, 6):
                    arr41.append(wb3.worksheets[0].cell(rowNum1,columnNum4).value)
                
                
                arr1.append(arr11)
                arr2.append(arr12)
                arr3.append(arr13)
                arr4.append(arr41)
            print(arr4)


        except Exception:

            # 第一步，创建一个logger
            logger = logging.getLogger()
            logger.setLevel(logging.INFO)  # Log等级总开关
            # 第二步，创建一个handler，用于写入日志文件
            rq = time.strftime('%Y%m%d%H%M', time.localtime(time.time()))
            log_path = './Logs/'
            log_name = log_path + rq + '.log'
            logfile = log_name
            fh = logging.FileHandler(logfile, mode='a')
            fh.setLevel(logging.DEBUG)  # 输出到file的log等级的开关
            # 第三步，定义handler的输出格式
            formatter = logging.Formatter(
                "%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s"
            )
            fh.setFormatter(formatter)
            # 第四步，将logger添加到handler里面
            logger.addHandler(fh)
            logger.error("Faild to open sklearn.txt from logger.error",
                         exc_info=True)


if (os.path.exists('./Logs') == False):
    os.mkdir("Logs")

update(a)
end = time.perf_counter()
print(end - start)